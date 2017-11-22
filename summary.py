'''


@author: gjc1211
'''
import xlrd
import xlwt
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from math import ceil

class Excel_Summary(object):
    
    
    def output_summary(self,Dist_Mark):
        HP_data = xlrd.open_workbook('HP_output.xlsx')
   #     print(HP_data)

        raw_data = []
        local_time_A = time.strftime('%Y-%m-%d',time.localtime(time.time()))
        time_sum = int(time.strftime('%Y',time.localtime(time.time()))) + int(time.strftime('%m',time.localtime(time.time()))) + int(time.strftime('%d',time.localtime(time.time())))
     #   print(time_sum)
   #     print(type(time_sum))
        data = xlrd.open_workbook('HP_output.xlsx')   #
        table = data.sheets()[0] #
        nrows = table.nrows
   #     print(nrows) 
        raw_data = (table.row_values(2)[3])
        print(raw_data)
        
        ex=load_workbook(filename=r'summary.xlsx')
        print('open excel success!')
        ws = ex.get_sheet_by_name("summary")
        print('open sheet1 success!')
        
        time_ref = "2017-11-16 00:00:00"
        timeArray_ref = time.strptime(time_ref, "%Y-%m-%d %H:%M:%S")
        local_time = int(time.time())/3600
        timeStamp_ref =  int(time.mktime(timeArray_ref))/3600
        time_inc = ceil((local_time - timeStamp_ref)/24)
        
        
        ws.cell(row=time_inc+1, column=2).value = local_time_A
        ws.cell(row=time_inc+1, column=Dist_Mark+2).value = raw_data
        print('write values success!')
        ex.save(filename='summary.xlsx')
        print('save success!')
        
#         wb = xlrd.open_workbook('summary.xlsx')
#         ws = wb.sheet_by_index(0)  
# 
#         ws.write(time_sum-2043, 1, local_time_A)
#         ws.write(time_sum-2043, 2, raw_data)
 
#        wb.save('summary.xlsx')